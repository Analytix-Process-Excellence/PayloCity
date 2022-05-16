import os.path, csv, datetime
import re
import time
from datetime import timedelta
from openpyxl import load_workbook,Workbook

class Paylocity_Process:

    def __init__(self,gui_queue,entrydate,filepath,narration):
        self.gui_queue = gui_queue
        self.entry_date = entrydate
        self.filepath = filepath
        self.narration = narration
        self.header = ['*Narration','*Date','Description','*AccountCode','*TaxRate','*Amount','TrackingName1','TrackingOption1','TrackingName2','TrackingOption2']
        self.xldata = {}
        self.tax = 'Tax Exempt (0%)'

    def pdftoexcel(self,entry,colname,amtcol):
        pdfextracterPath = "pdftextExtracter.exe"
        pdfpath = os.path.join(self.filepath,'Payroll Register.pdf')
        myCmd = pdfextracterPath + f" -layout \"{pdfpath}\""
        os.system(myCmd)
        time.sleep(1)
        textName = str(pdfpath).replace('pdf', 'txt').replace('PDF', 'txt')
        textName = os.path.abspath(textName)

        if not textName:
            msg = f'Unable to convert PDF file.'
            self.gui_queue.put({'status': msg})
            return False
        save_xl_Name = str(pdfpath).replace('pdf', 'xlsx').replace('PDF', 'xlsx')
        save_xl_Name = os.path.abspath(save_xl_Name)
        if os.path.isfile(save_xl_Name):
            return save_xl_Name

        created_xl = Workbook()
        created_xs = created_xl.active

        with open(textName, 'r') as txt:
            for line in txt:
                x = re.split(r'\s\s+', line.replace('\f', ''))
                try:
                    if len(x) > 0:
                        created_xs.append(x)
                except:
                    print('unable to print')
        amt = 0.00
        created_xl.save(save_xl_Name)
        os.remove(textName)
        flag = False
        for rows in created_xs.values:
            if 'report totals' in str(rows[0]).strip().lower():
                flag = True
            elif flag:
                for num,row in enumerate(rows):
                    if row:
                        if entry in row and '/' not in row:
                            if rows[amtcol]:
                                amt = float(str(rows[amtcol]).replace(',','').strip())
                            break
        os.remove(save_xl_Name)
        return amt


    def csvtoexcel(self,laborname):
        laborwb = Workbook()
        laborws = laborwb.active
        with open(laborname) as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                laborws.append(row)
        return laborws


    def process(self,entry,glcode,desc,multi,colname,amtcol):
        laborname = os.path.join(self.filepath,'Labor Distribution Data Export.csv')

        laborws = self.csvtoexcel(laborname)
        for labor in laborws.values:
            if labor:
                if desc == 'Wages':
                    if entry == str(labor[colname]).split('-')[1] and "Earnings" in labor[5] and labor[6] != "PDTPS" and labor[6] != "PRLTP" and labor[amtcol]:
                        if self.xldata:
                            if (
                            self.narration, self.entry_date, desc, entry, glcode, self.tax,labor[6]) in self.xldata.keys():
                                self.xldata[self.narration, self.entry_date, desc, entry, glcode, self.tax,labor[6]] += float(labor[amtcol]) * multi

                            elif (self.narration, self.entry_date, desc, entry, glcode,
                                  self.tax,labor[6]) not in self.xldata.keys():
                                self.xldata[self.narration, self.entry_date, desc, entry, glcode, self.tax,labor[6]] = float(labor[amtcol]) * multi
                        else:
                            self.xldata[self.narration, self.entry_date, desc, entry, glcode, self.tax,labor[6]] = float(labor[amtcol]) * multi
                elif desc == "Tips":
                    if entry == labor[colname] and "Earnings" in labor[5] and labor[amtcol]:
                        if self.xldata:
                            if (
                            self.narration, self.entry_date, desc, entry, glcode, self.tax,labor[6]) in self.xldata.keys():
                                self.xldata[self.narration, self.entry_date, desc, entry, glcode, self.tax,labor[6]] += float(labor[amtcol]) * multi

                            elif (self.narration, self.entry_date, desc, entry, glcode,
                                  self.tax,labor[6]) not in self.xldata.keys():
                                self.xldata[self.narration, self.entry_date, desc, entry, glcode, self.tax,labor[6]] = float(labor[amtcol]) * multi
                        else:
                            self.xldata[self.narration, self.entry_date, desc, entry, glcode, self.tax,labor[6]] = float(labor[amtcol]) * multi
                elif desc == "Taxes":
                    if entry == labor[colname] and desc in labor[5] and labor[amtcol]:
                        if self.xldata:
                            if (
                            self.narration, self.entry_date, desc, entry, glcode, self.tax,labor[6]) in self.xldata.keys():
                                self.xldata[self.narration, self.entry_date, desc, entry, glcode, self.tax,labor[6]] += float(labor[amtcol]) * multi

                            elif (self.narration, self.entry_date, desc, entry, glcode,
                                  self.tax,labor[6]) not in self.xldata.keys():
                                self.xldata[self.narration, self.entry_date, desc, entry, glcode, self.tax,labor[6]] = float(labor[amtcol]) * multi
                        else:
                            self.xldata[self.narration, self.entry_date, desc, entry, glcode, self.tax,labor[6]] = float(labor[amtcol]) * multi
                elif desc == "Deductions":
                    if entry == labor[colname] and desc in labor[5] and labor[amtcol]:
                        if self.xldata:
                            if (
                                    self.narration, self.entry_date, desc, entry, glcode, self.tax,
                                    labor[6]) in self.xldata.keys():
                                self.xldata[
                                    self.narration, self.entry_date, desc, entry, glcode, self.tax, labor[6]] += float(
                                    labor[amtcol]) * multi

                            elif (self.narration, self.entry_date, desc, entry, glcode,
                                  self.tax, labor[6]) not in self.xldata.keys():
                                self.xldata[
                                    self.narration, self.entry_date, desc, entry, glcode, self.tax, labor[6]] = float(
                                    labor[amtcol]) * multi
                        else:
                            self.xldata[
                                self.narration, self.entry_date, desc, entry, glcode, self.tax, labor[6]] = float(
                                labor[amtcol]) * multi
                elif desc == "Employer Taxes":
                    if entry == labor[colname] and desc in labor[5] and labor[amtcol]:
                        if self.xldata:
                            if (
                                    self.narration, self.entry_date, desc, entry, glcode, self.tax,
                                    labor[6]) in self.xldata.keys():
                                self.xldata[
                                    self.narration, self.entry_date, desc, entry, glcode, self.tax, labor[6]] += float(
                                    labor[amtcol]) * multi

                            elif (self.narration, self.entry_date, desc, entry, glcode,
                                  self.tax, labor[6]) not in self.xldata.keys():
                                self.xldata[
                                    self.narration, self.entry_date, desc, entry, glcode, self.tax, labor[6]] = float(
                                    labor[amtcol]) * multi
                        else:
                            self.xldata[
                                self.narration, self.entry_date, desc, entry, glcode, self.tax, labor[6]] = float(
                                labor[amtcol]) * multi
                        if (self.narration, self.entry_date, 'Taxes', None, 2235, self.tax, None) in self.xldata.keys():
                            self.xldata[
                                self.narration, self.entry_date, 'Taxes', None, 2235, self.tax, None] += float(
                                labor[amtcol]) * -1
                        else:
                            self.xldata[
                                self.narration, self.entry_date, 'Taxes', None, 2235, self.tax, None] = float(
                                labor[amtcol]) * -1
                elif desc == "DD":
                    amt = self.pdftoexcel(entry,colname,amtcol)
                    self.xldata[
                        self.narration, self.entry_date, desc, entry, glcode, self.tax, labor[6]] = float(
                        amt) * multi
                    break
                elif desc == "Checks":
                    amt = self.pdftoexcel(entry,colname,amtcol)
                    self.xldata[
                        self.narration, self.entry_date, desc, entry, glcode, self.tax, labor[6]] = float(
                        amt) * multi
                    break
        return True

    def xlupdate(self,entrydate):
        entrydate = entrydate.replace('/','-')
        filename = os.path.join(os.getcwd(),entrydate)
        if not os.path.exists(filename):
            os.mkdir(filename)
        filename = os.path.join(filename,'JournalEntry.xlsx')
        journalwb = Workbook()
        journalws = journalwb.active
        journalws.title = "JournalEntry"
        journalws.append(self.header)
        for key, value in self.xldata.items():
            if value != 0:
                journalws.append([key[0],key[1],key[3],key[4],key[5],value,None,None,None])
        journalwb.save(filename)
        for row in journalws.iter_rows(min_row=2, min_col=6, max_col=6):
            for cell in row:
                cell.number_format = f'#,##0.00'
        journalwb.save(filename)
        return True

class runPay:

    def __init__(self):
        self.gui_queue = None

    def run(self,filepath,startdate,enddate):
        run_start = time.perf_counter()
        self.gui_queue.put({'status': f'Paylocity Journal Proces started...'}) if self.gui_queue else None
        settingsheet = 'PaylocitySettingSheet.xlsx'
        settingwb = load_workbook(settingsheet,read_only=True, data_only=True)
        settingws = settingwb['EntryDetails']
        entrydate = enddate.strftime("%m/%d/%Y")
        narration = f'Payroll WE_{entrydate}'
        entrydate = enddate.strftime("%#m/%d/%Y")
        paylo = Paylocity_Process(self.gui_queue,entrydate,filepath,narration)
        for num,settingdata in enumerate(settingws.values):
            if num == 0:
                continue
            entry = settingdata[0]
            glcode = settingdata[1]
            desc = settingdata[2]
            multi = settingdata[3]
            colname = settingdata[4]-1
            amtcol = settingdata[5]-1
            paystatus = paylo.process(entry,glcode,desc,multi,colname,amtcol)
            if not paystatus:
                self.gui_queue.put({'status':f'Error in entry {entry}'}) if self.gui_queue else None
        xlupdate = paylo.xlupdate(entrydate)
        if not xlupdate:
            self.gui_queue.put({'status': f'Error in creating journal'}) if self.gui_queue else None

        self.gui_queue.put({'status': f'Paylocity Journal Proces completed'}) if self.gui_queue else None
        run_end = time.perf_counter()
        time_taken = time.strftime("%H:%M:%S", time.gmtime(int(run_end - run_start)))
        self.gui_queue.put({"status": f"Time Taken {time_taken}"})
        return True